import pandas as pd
import numpy as np
import os, sys
sys.path.insert(0, "../")
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
import yaml

from factor_cal.config_loader import basic_config as cfg
from factor_cal.utils import ddb_utils as du
from factor_cal.factor_eval.basic_evaluate import get_factor_ic_summary_info
from factor_cal.feature import features as fe
from factor_cal.factor import factors as fa

class Condition:
    def __init__(self, name, value, func):
        self._name = name
        self._value = value
        self._func = func
        
    @property
    def name(self):
        return self._name
    
    @property
    def value(self):
        return self._value
    
    @value.setter
    def value(self, value):
        self._value = value
    
    @property
    def index(self):
        return self._index
    
    @index.setter
    def index(self, index):
        self._index = index
    
    def is_satisfied(self, value):
        return self._func(value, self._value)

def abs_cond(value, threshold):
    return abs(value) > threshold

def win_rate_cond(value, threshold):
    return abs(value-0.5) > threshold

def determine_index(dic, conditions, name, index):
    for condition in conditions:
        if name == condition._name:
            condition.index = index
            dic[index] = condition


def process_perFactor(factor_name, start_date, end_date, config, base_dir):
    csv_dir = f'{base_dir}/tsICtable/{factor_name}'
    
    summary_results = []
    ic_results = []
    for date in pd.date_range(start_date, end_date):
        date = date.strftime('%Y.%m.%d')
        ic_filepath = f'{csv_dir}/{date}_byCode.csv'
        if not os.path.exists(ic_filepath):
            continue
        ic_result = pd.read_csv(ic_filepath)
        
        tmp_result = ic_result.copy()
        tmp_result['date'] = date
        ic_results.append(tmp_result) 
        
        ic_result.set_index(['securityid'], inplace=True)
        ic_summary = get_factor_ic_summary_info(ic_result)
        ic_summary['IC effective rate'] = (ic_result.abs() > 0.03).sum() / ic_result.count()
        
        ic_summary = pd.DataFrame(ic_summary.loc['1m'].to_dict(), index=[date])
        summary_results.append(ic_summary)    
    
    result_df = pd.concat(summary_results, axis=0)
    result_df.index.set_names(['Date'], inplace=True)
    result_df.reset_index(inplace=True)
    save_filepath = f'{csv_dir}/summary.csv'
    result_df.to_csv(save_filepath)
    
    plot_dir = f"{base_dir}/tsICPlot/{factor_name}"
    os.makedirs(plot_dir, exist_ok=True)
    plot_filepath = f"{plot_dir}/IC_and_effective_rate.png"
    plot_summary(result_df, factor_name, plot_filepath)
    
    ic_total_df = pd.concat(ic_results, axis=0)
    ic_total_df.set_index(['date', 'securityid'], inplace=True)
    ic_total_summary = get_factor_ic_summary_info(ic_total_df)
    factor_result = ic_total_summary.loc['1m', ['IC Mean', 'ICIR']]
    
    return pd.DataFrame(factor_result).T
    
def plot_summary(result_df, factor_name, save_filepath):
    # create figure
    fig, ax1 = plt.subplots()

    ax1.bar(result_df['Date'], result_df['IC Mean'], color='gray', alpha=0.6, label='IC')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('IC', color='gray')
    ax1.tick_params(axis='y', labelcolor='gray')

    ax2 = ax1.twinx()
    ax2.plot(result_df['Date'], result_df['IC effective rate'], color='blue', label='Eff. rate')
    ax2.set_ylabel('Effective rate', color='blue')
    ax2.tick_params(axis='y', labelcolor='blue')
    ax2.set_ylim(bottom=0, top=1)


    # 添加图例
    fig.legend()
    # fig.legend(loc='upper left', bbox_to_anchor=(0.1, 0.9))

    # 显示图形
    plt.title(f'Factor: {factor_name}')
    fig.savefig(save_filepath)
    plt.close()

def get_best_parameter(factor_name):
    base_dirpath = "/home/wangzirui/workspace/data/param_scan"
    param_filepath = f"{base_dirpath}/{factor_name}/2023.09.22.pkl"
    if not os.path.exists(param_filepath):
        return None
    df = pd.read_pickle(param_filepath)
    idx = df['metric'].argmax()
    param = df.iloc[idx].drop('metric').to_dict()
    return param

def process_xlsx_file(xlsx_filepath, base_dir):

    book = load_workbook(xlsx_filepath)
    sheet = book['Sheet1']
    
    conditions = [Condition('IC Mean', 0.03, abs_cond), Condition('ICIR', 1.5, abs_cond)]
     
    fac_index = None
    satisfied_factors = []
    cond_ind_dict = {}
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            for j, cell in enumerate(row):
                if cell.value == 'factor':
                    fac_index = j
                    continue
                determine_index(cond_ind_dict, conditions, cell.value, j)
        else:
            satisfied_cond_num = 0
            factor_name = None
            for j, cell in enumerate(row):
                if j == fac_index:
                    factor_name = cell.value
                    continue
                if j in cond_ind_dict.keys():
                    condition = cond_ind_dict[j]
                    if condition.is_satisfied(cell.value):
                        satisfied_cond_num += 1
                        cell.font = Font(color='FF0000')
                
            if satisfied_cond_num >= 1:
                satisfied_factors.append(factor_name)
            
                         
    book.save(xlsx_filepath)
    
    yaml.dump(satisfied_factors, open(os.path.join(base_dir, f'satisfied_factors.yml'), 'w'))


if __name__ == "__main__":
    # read config file
    config = cfg.BasicConfig('config/config_scan.yml')
    
    features = fe.Features(config)
    factors = fa.Factors(config, features)
    
    start_date = '2023.09.22'
    end_date = '2023.09.30'
    base_dir = '/home/wangzirui/workspace/factor_eval_summary/param_optimized'
    
    ic_results = None
    for facType in config['factors']:
        for i, factor_name in enumerate(config['factors'][facType]):
            # if (i >= 1):
            #     break
            print("Evaluating factor: ", factor_name, " ...", flush=True)
            factor_result = process_perFactor(factor_name, start_date, end_date, config, base_dir)
            factor = factors.fac_dict.get(factor_name)
            factor_result['factor'] = factor_name
            parameter = get_best_parameter(factor_name)
            factor_result['formula'] = str(factor)
            factor_result['best_param'] = str(parameter)
            factor_result['description'] = factor.desc
            ic_results = pd.concat([ic_results, factor_result], axis=0, ignore_index=True)
    
    xlsx_filepath = f"{base_dir}/ic_summary.xlsx"
    ic_results.to_excel(xlsx_filepath, index=False)
    
    process_xlsx_file(xlsx_filepath, base_dir)
    